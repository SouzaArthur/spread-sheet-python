import openpyxl
import os
from openpyxl.utils import get_column_letter

os.chdir('/Users/souzaarthur/workspace/pessoal/spreadSheet/workBooks/')
wb = openpyxl.load_workbook('spreadSheetTest.xlsx')
print(type(wb))
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet)

print("sheet title")
print(sheet.title)

# Here I get the values from cell

c = sheet['B1']

print(sheet['B2'].value)

toPrint = 'Row %s, Column %s is %s' %(c.row, c.column, c.value)

print(toPrint)

print(c.coordinate)

for i in range(1, 8, 1):
    print(i, sheet.cell(row=i, column=2).value)

print("max column and row")

print(sheet.max_column)
print(sheet.max_row)

# Every cell printed

lastCellAndRow = get_column_letter(sheet.max_column) + str(sheet.max_row)

print(lastCellAndRow)
for objectsInARange in sheet['A1':lastCellAndRow]:
    for cell in objectsInARange:
        print(cell.coordinate, cell.value)
    print("-------------------")

# Showing every fruit

sheet['D1'] = "Hello, World!"

wb.save('spreadSheetTest.xlsx')